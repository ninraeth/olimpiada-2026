#!/usr/bin/env python3
"""Export local xlsx to data/sample.json using the same section logic (smoke test)."""

import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "Olimpiada2026_v2.xlsx"
OUT = ROOT / "data" / "sample.json"


def rows_from_sheet(ws):
    out = []
    for row in ws.iter_rows(max_row=ws.max_row or 1, max_col=ws.max_column or 1, values_only=True):
        out.append(["" if v is None else str(v).strip() for v in row])
    return out


def strip_leading(rows):
    if not rows:
        return rows
    min_lead = 10**9
    for row in rows:
        lead = 0
        while lead < len(row) and not row[lead]:
            lead += 1
        if lead < len(row):
            min_lead = min(min_lead, lead)
    if min_lead == 10**9 or min_lead == 0:
        return rows
    return [r[min_lead:] for r in rows]


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    payload = {"sheets": {}}
    for name in wb.sheetnames:
        rows = strip_leading(rows_from_sheet(wb[name]))
        # drop fully empty trailing cols
        payload["sheets"][name] = rows
        markers = [r[0] for r in rows if r and r[0].startswith("#")]
        print(f"{name}: markers={markers}")
    OUT.parent.mkdir(exist_ok=True)
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
