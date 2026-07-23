#!/usr/bin/env python3
"""
Append # STREFA MEDALOWA to each discipline sheet in the Google workbook.

Downloads the live xlsx, adds medal sections where missing, and tries to
write back. Without OAuth the write step is skipped and a local file is saved
for manual upload.
"""
from __future__ import annotations

import io
import re
import urllib.request
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ID = "18Frm47PTR0FCaZs4QoELydkQNmLQWmvU"
DISCIPLINE_SHEETS = {
    "Piłka Nożna",
    "Siatkówka",
    "Koszykówka",
    "Piłka ind.",
    "Badminton",
    "Inne",
}

HEADER = ["medal", "nazwa", "gracze"]
ROWS = [
    ["złoty", "", ""],
    ["srebrny", "", ""],
    ["brązowy", "", ""],
]

section_font = Font(name="Calibri", size=12, bold=True, color="1A365D")
header_font = Font(name="Calibri", size=11, bold=True)
header_fill = PatternFill("solid", fgColor="E2E8F0")
yellow_fill = PatternFill("solid", fgColor="FEFCBF")
gold_fill = PatternFill("solid", fgColor="FEF3C7")
silver_fill = PatternFill("solid", fgColor="E2E8F0")
bronze_fill = PatternFill("solid", fgColor="FFEDD5")
thin = Border(
    left=Side(style="thin", color="CBD5E0"),
    right=Side(style="thin", color="CBD5E0"),
    top=Side(style="thin", color="CBD5E0"),
    bottom=Side(style="thin", color="CBD5E0"),
)
note_font = Font(name="Calibri", size=10, italic=True, color="718096")


def download_xlsx() -> bytes:
    url = f"https://docs.google.com/spreadsheets/d/{ID}/export?format=xlsx"
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    return urllib.request.urlopen(req, timeout=60).read()


def sheet_has_medal_zone(ws) -> bool:
    for row in ws.iter_rows(max_row=ws.max_row or 1, max_col=3, values_only=True):
        for cell in row:
            if cell is None:
                continue
            s = str(cell)
            if re.search(r"#\s*STREFA\s*MEDALOWA|#\s*MEDALE", s, re.I):
                return True
    return False


def append_medal_zone(ws) -> None:
    # Find last used row
    last = ws.max_row or 1
    while last > 1:
        vals = [ws.cell(last, c).value for c in range(1, 8)]
        if any(v is not None and str(v).strip() for v in vals):
            break
        last -= 1

    r = last + 2
    ws.cell(r, 1, "# STREFA MEDALOWA").font = section_font
    r += 1
    for i, h in enumerate(HEADER, 1):
        cell = ws.cell(r, i, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin
    r += 1
    fills = [gold_fill, silver_fill, bronze_fill]
    for i, row in enumerate(ROWS):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            cell.border = thin
            if c == 1:
                cell.fill = fills[i]
                cell.font = header_font
            else:
                cell.fill = yellow_fill
        r += 1
    ws.cell(
        r + 1,
        1,
        "* Wpisz ręcznie: medal (złoty/srebrny/brązowy), nazwa gracza lub drużyny, "
        "opcjonalnie skład drużyny (gracze po przecinku).",
    ).font = note_font


def main():
    print("Downloading workbook…")
    raw = download_xlsx()
    wb = load_workbook(io.BytesIO(raw))
    print("Sheets:", wb.sheetnames)

    changed = []
    for name in wb.sheetnames:
        if name not in DISCIPLINE_SHEETS:
            print(f"  skip {name!r}")
            continue
        ws = wb[name]
        if sheet_has_medal_zone(ws):
            print(f"  already has medal zone: {name!r}")
            continue
        append_medal_zone(ws)
        changed.append(name)
        print(f"  added medal zone: {name!r}")

    out = Path("Olimpiada2026_medals.xlsx")
    wb.save(out)
    print(f"Saved local copy: {out.resolve()}")
    print(
        "\nAby zaktualizować Google Sheets:\n"
        "1) Otwórz dokument Olimpiady\n"
        "2) Plik → Importuj → Prześlij → Olimpiada2026_medals.xlsx\n"
        "   LUB ręcznie wklej na końcu każdej dyscypliny blok:\n"
        "   # STREFA MEDALOWA\n"
        "   medal | nazwa | gracze\n"
        "   złoty | | \n"
        "   srebrny | | \n"
        "   brązowy | | \n"
    )
    if changed:
        print("Zmienione arkusze:", ", ".join(changed))
    else:
        print("Brak nowych sekcji do dodania (już istnieją).")


if __name__ == "__main__":
    main()
