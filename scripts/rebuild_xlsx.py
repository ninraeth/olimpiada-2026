#!/usr/bin/env python3
"""Rebuild Olimpiada2026_v2.xlsx with machine-friendly section markers."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

title_font = Font(name="Calibri", size=18, bold=True, color="1A365D")
section_font = Font(name="Calibri", size=12, bold=True, color="1A365D")
header_font = Font(name="Calibri", size=11, bold=True)
header_fill = PatternFill("solid", fgColor="E2E8F0")
id_fill = PatternFill("solid", fgColor="CBD5E0")
yellow_fill = PatternFill("solid", fgColor="FEFCBF")
green_fill = PatternFill("solid", fgColor="C6F6D5")
note_font = Font(name="Calibri", size=10, italic=True, color="718096")
thin = Border(
    left=Side(style="thin", color="CBD5E0"),
    right=Side(style="thin", color="CBD5E0"),
    top=Side(style="thin", color="CBD5E0"),
    bottom=Side(style="thin", color="CBD5E0"),
)

TEAMS = [
    (1, "Drużyna 1", "Jan Kowalski, Piotr Nowak, Adam Wiśniewski"),
    (2, "Drużyna 2", "Marek Zieliński, Tomasz Lewandowski, Krzysztof Wójcik"),
    (3, "Drużyna 3", "Paweł Kamiński, Michał Szymański, Andrzej Woźniak"),
    (4, "Drużyna 4", "Robert Kozłowski, Łukasz Jankowski, Marcin Mazur"),
]


def style_header_row(ws, row, cols, id_cols=None):
    id_cols = id_cols or set()
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = id_fill if c in id_cols else header_fill
        cell.border = thin
        cell.alignment = Alignment(wrap_text=True)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_team_discipline(ws, title, score_note, ranking_headers, ranking_notes):
    ws["A1"] = title
    ws["A1"].font = title_font

    ws["A3"] = "# DRUŻYNY"
    ws["A3"].font = section_font
    ws["A4"] = "ID_drużyny"
    ws["B4"] = "Nazwa drużyny"
    ws["C4"] = "Gracze (oddzieleni przecinkiem)"
    style_header_row(ws, 4, 3, id_cols={1})
    for i, (tid, name, players) in enumerate(TEAMS):
        r = 5 + i
        ws.cell(r, 1, tid).fill = id_fill
        ws.cell(r, 2, name).fill = yellow_fill
        ws.cell(r, 3, players).fill = yellow_fill
        for c in range(1, 4):
            ws.cell(r, c).border = thin

    ws["A10"] = "# MECZE"
    ws["A10"].font = section_font
    headers = ["ID_meczu", "Faza", "Drużyna 1", "Drużyna 2", "Wynik (X:Y)"]
    for c, h in enumerate(headers, 1):
        ws.cell(11, c, h)
    style_header_row(ws, 11, 5, id_cols={1})
    matches = [
        (1, "1/2 Finału", "Drużyna 1", "Drużyna 2", ""),
        (2, "1/2 Finału", "Drużyna 3", "Drużyna 4", ""),
        (3, "Mecz o 3. miejsce", "", "", ""),
        (4, "Finał", "", "", ""),
    ]
    for i, (mid, phase, d1, d2, score) in enumerate(matches):
        r = 12 + i
        ws.cell(r, 1, mid).fill = id_fill
        ws.cell(r, 2, phase).fill = yellow_fill
        ws.cell(r, 3, d1).fill = yellow_fill
        ws.cell(r, 4, d2).fill = yellow_fill
        ws.cell(r, 5, score).fill = green_fill
        for c in range(1, 6):
            ws.cell(r, c).border = thin

    ws["A17"] = score_note
    ws["A17"].font = note_font

    ws["A19"] = "# RANKING"
    ws["A19"].font = section_font
    for c, h in enumerate(ranking_headers, 1):
        ws.cell(20, c, h)
    style_header_row(ws, 20, len(ranking_headers))
    for i, note in enumerate(ranking_notes):
        ws.cell(22 + i, 1, note).font = note_font
    set_col_widths(ws, [14, 22, 50, 18, 14])


# ========== INFO ==========
ws = wb.active
ws.title = "Info"
ws["A1"] = "Olimpiada Bieździadów 2026"
ws["A1"].font = title_font
ws["A3"] = "Informacje o rozgrywkach"
ws["A3"].font = section_font
ws["A5"] = "Dyscypliny:"
ws["A5"].font = header_font
ws["A6"] = "• Piłka Nożna (drużynowa)"
ws["A7"] = "• Siatkówka (drużynowa)"
ws["A8"] = "• Koszykówka (indywidualna - rzuty)"
ws["A9"] = "• Badminton (indywidualny)"
ws["A11"] = "Uwagi ogólne:"
ws["A11"].font = header_font
ws["A12"] = "1. Format wyniku we wszystkich meczach 1v1 / drużyna vs drużyna: X:Y"
ws["A13"] = "2. Kolumny oznaczone szarym kolorem (ID_...) nie powinny być wyświetlane w aplikacji."
ws["A14"] = "3. Żółte pola = pola do wpisywania danych (edytowalne)."
ws["A15"] = "4. Zielone pola = wyniki."
ws["A16"] = "5. Rankingi indywidualne powinny być sortowane według reguł podanych w każdej zakładce."
ws["A17"] = "6. Sekcje w arkuszach oznaczane są znacznikami # DRUŻYNY, # MECZE, # RANKING, # GRACZE, # SEKCJA."
ws["A18"] = "7. Aby dodać mecz lub gracza — dopisz nowy wiersz pod tabelą (nie zostawiaj pustych ID-only)."
ws["A20"] = "data"
ws["B20"] = ""
ws["A20"].fill = yellow_fill
ws["B20"].fill = yellow_fill
ws["A21"] = "miejsce"
ws["B21"] = ""
ws["A21"].fill = yellow_fill
ws["B21"].fill = yellow_fill
ws["A23"] = "Wersja szablonu: 2.1"
ws["A23"].font = note_font
set_col_widths(ws, [80, 40])

# ========== PILKA / SIATKOWKA ==========
ws = wb.create_sheet("Piłka Nożna")
build_team_discipline(
    ws,
    "Piłka Nożna",
    "* Format wyniku: X:Y (X = gole Drużyny 1, Y = gole Drużyny 2)",
    ["miejsce", "gracz", "zwycięstwa / mecze (%)", "różnica goli", "uwagi"],
    [
        "* Sortowanie: najpierw po zwycięstwa/mecze (%) (malejąco), przy remisie po różnicy goli (malejąco).",
        "* Różnica goli = gole strzelone przez drużyny gracza minus gole stracone.",
        "* Ranking wypełniany ręcznie. Aplikacja pokazuje tylko wiersze z wypełnionym graczem.",
    ],
)

ws = wb.create_sheet("Siatkówka")
build_team_discipline(
    ws,
    "Siatkówka",
    "* Format wyniku: X:Y (X = wygrane sety Drużyny 1, Y = wygrane sety Drużyny 2)",
    ["miejsce", "gracz", "zwycięstwa / mecze (%)", "różnica setów", "uwagi"],
    [
        "* Sortowanie: najpierw po zwycięstwa/mecze (%) (malejąco), przy remisie po różnicy setów (malejąco).",
        "* Różnica setów = sety wygrane przez drużyny gracza minus sety przegrane.",
        "* Ranking wypełniany ręcznie. Aplikacja pokazuje tylko wiersze z wypełnionym graczem.",
    ],
)

# ========== KOSZYKOWKA ==========
ws = wb.create_sheet("Koszykówka")
ws["A1"] = "Koszykówka (rzuty)"
ws["A1"].font = title_font
ws["A3"] = "# GRACZE"
ws["A3"].font = section_font
headers = [
    "ID_gracza",
    "Imię gracza",
    "WYNIK",
    "Próba 1 - 1P",
    "Próba 1 - 2P",
    "Próba 1 - 3P",
    "Próba 1 - UK1",
    "Próba 1 - UK2",
    "Próba 2 - 1P",
    "Próba 2 - 2P",
    "Próba 2 - 3P",
    "Próba 2 - UK1",
    "Próba 2 - UK2",
    "Próba 3 - 1P",
    "Próba 3 - 2P",
    "Próba 3 - 3P",
    "Próba 3 - UK1",
    "Próba 3 - UK2",
]
for c, h in enumerate(headers, 1):
    ws.cell(4, c, h)
style_header_row(ws, 4, 18, id_cols={1})
ws.cell(4, 3).fill = green_fill
for i in range(1, 13):
    r = 4 + i
    ws.cell(r, 1, i).fill = id_fill
    ws.cell(r, 2, f"Gracz {i}").fill = yellow_fill
    formula = f'=IFERROR(AVERAGE(D{r}:R{r}),"")'
    ws.cell(r, 3, formula).fill = green_fill
    for c in range(4, 19):
        ws.cell(r, c).fill = yellow_fill
        ws.cell(r, c).border = thin
    for c in range(1, 4):
        ws.cell(r, c).border = thin
ws["A18"] = "* WYNIK = średnia ze wszystkich wypełnionych pól prób. Sortowanie w aplikacji po WYNIK malejąco."
ws["A18"].font = note_font
ws["A19"] = "* Kolumna ID_gracza (szara) — nie wyświetlać w aplikacji."
ws["A19"].font = note_font
ws["A20"] = "* 1P / 2P / 3P / UK1 / UK2 = punkty z danej próby (rzuty)."
ws["A20"].font = note_font
set_col_widths(ws, [12, 16, 10] + [12] * 15)

# ========== BADMINTON ==========
ws = wb.create_sheet("Badminton")
ws["A1"] = "Badminton"
ws["A1"].font = title_font
ws["A3"] = "# GRACZE"
ws["A3"].font = section_font
ws["A4"] = "ID_gracza"
ws["B4"] = "Imię gracza"
style_header_row(ws, 4, 2, id_cols={1})
for i in range(1, 7):
    r = 4 + i
    ws.cell(r, 1, i).fill = id_fill
    ws.cell(r, 2, f"Gracz {i}").fill = yellow_fill
    ws.cell(r, 1).border = thin
    ws.cell(r, 2).border = thin

ws["A12"] = "# MECZE"
ws["A12"].font = section_font
for c, h in enumerate(["ID_meczu", "Faza", "Gracz 1", "Gracz 2", "Wynik (X:Y)"], 1):
    ws.cell(13, c, h)
style_header_row(ws, 13, 5, id_cols={1})
badminton_matches = [
    (1, "Eliminacje", "Gracz 1", "Gracz 2", ""),
    (2, "Eliminacje", "Gracz 3", "Gracz 4", ""),
    (3, "Eliminacje", "Gracz 5", "Gracz 6", ""),
    (4, "1/4 Finału", "", "", ""),
    (5, "1/4 Finału", "", "", ""),
    (6, "1/4 Finału", "", "", ""),
    (7, "1/4 Finału", "", "", ""),
    (8, "1/2 Finału", "", "", ""),
    (9, "1/2 Finału", "", "", ""),
    (10, "Mecz o 3. miejsce", "", "", ""),
    (11, "Finał", "", "", ""),
]
for i, (mid, phase, g1, g2, score) in enumerate(badminton_matches):
    r = 14 + i
    ws.cell(r, 1, mid).fill = id_fill
    ws.cell(r, 2, phase).fill = yellow_fill
    ws.cell(r, 3, g1).fill = yellow_fill
    ws.cell(r, 4, g2).fill = yellow_fill
    ws.cell(r, 5, score).fill = green_fill
    for c in range(1, 6):
        ws.cell(r, c).border = thin

ws["A27"] = "# RANKING"
ws["A27"].font = section_font
for c, h in enumerate(["miejsce", "gracz", "zwycięstwa / mecze (%)", "uwagi"], 1):
    ws.cell(28, c, h)
style_header_row(ws, 28, 4)
ws["A30"] = "* Format wyniku: X:Y (X = sety/punkty Gracza 1, Y = sety/punkty Gracza 2)"
ws["A30"].font = note_font
ws["A31"] = "* Jeśli brak listy GRACZE — aplikacja zbierze imiona z meczów."
ws["A31"].font = note_font
set_col_widths(ws, [12, 22, 16, 16, 14])

# ========== INNE ==========
ws = wb.create_sheet("Inne")
ws["A1"] = "Inne konkurencje"
ws["A1"].font = title_font
ws["A3"] = "# SEKCJA | Przykładowa konkurencja"
ws["A3"].font = section_font
for c, h in enumerate(["miejsce", "uczestnik", "wynik", "uwagi"], 1):
    ws.cell(4, c, h)
style_header_row(ws, 4, 4)
ws["A5"] = 1
ws["B5"] = "Uczestnik 1"
ws["C5"] = ""
ws["D5"] = ""
for c in range(1, 5):
    ws.cell(5, c).fill = yellow_fill if c > 1 else id_fill
    ws.cell(5, c).border = thin

ws["A7"] = "# SEKCJA | Druga konkurencja (opcjonalnie)"
ws["A7"].font = section_font
for c, h in enumerate(["miejsce", "uczestnik", "wynik", "uwagi"], 1):
    ws.cell(8, c, h)
style_header_row(ws, 8, 4)
ws["A10"] = "* Każda sekcja zaczyna się od: # SEKCJA | Nazwa konkurencji"
ws["A10"].font = note_font
ws["A11"] = "* Nagłówek tabeli w następnym wierszu, potem wiersze danych."
ws["A11"].font = note_font
ws["A12"] = "* Możesz dodać dowolne kolumny — aplikacja wyświetli je generycznie."
ws["A12"].font = note_font
set_col_widths(ws, [12, 22, 14, 30])

out = "Olimpiada2026_v2.xlsx"
wb.save(out)
print(f"Saved {out}, sheets: {wb.sheetnames}")
